"""Vendor analysis page + per-strength/weakness "why?" analysis.

Page route is server-rendered aggregation — pure read-only over data
already in the DB, no LLM. The per-item reason analysis (triggered by
clicking a strength or weakness on the page) is in this file too, since
it's just another view of the same vendor model.
"""
from __future__ import annotations

from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession

from app.db import get_session
from app.services.vendor_reasons import (
    delete_card,
    extract_reasons,
    find_card_lookup,
    get_card,
    list_saved_cards,
    reanalyze_card,
    save_card,
    toggle_card_hidden,
    update_card_label,
)
from app.services.vendors import list_vendors
from app.templating import render

router = APIRouter()


@router.get("/vendors")
async def vendors_page(request: Request, session: AsyncSession = Depends(get_session)):
    vendors = await list_vendors(session)
    # Build a (vendor_key, category_lower, band) → card_id map so the
    # template can render a 📌 marker next to strengths/weaknesses the
    # user already saved. Lookup happens off a composite index so even
    # a 1k-card workspace stays fast.
    saved_lookup = await find_card_lookup(session)
    return render(
        request,
        "vendors.html",
        vendors=vendors,
        saved_lookup=saved_lookup,
    )


# ----------------------------------------------------------------------------
# Reasons extraction (live LLM)
# ----------------------------------------------------------------------------


@router.get("/api/vendor-reasons")
async def vendor_reasons_endpoint(
    request: Request,
    vendor_key: str = Query(..., min_length=1, max_length=100),
    category_name: str = Query(..., min_length=1, max_length=200),
    band: str = Query(..., pattern="^(positive|negative)$"),
    summary_lang: str = Query("en"),
    model: Optional[str] = None,
    force: bool = Query(False),
    session: AsyncSession = Depends(get_session),
):
    """Live extraction. In-memory cached for 30 min by default; pass
    force=true to bypass."""
    try:
        return await extract_reasons(
            session,
            vendor_key=vendor_key,
            category_name=category_name,
            band=band,
            summary_lang=summary_lang,
            model=model,
            force=force,
        )
    except RuntimeError as e:
        raise HTTPException(503, str(e))
    except ValueError as e:
        raise HTTPException(400, str(e))


# ----------------------------------------------------------------------------
# Saved-card CRUD
# ----------------------------------------------------------------------------


class SaveReasonCardBody(BaseModel):
    label: Optional[str] = Field(None, max_length=200)
    # Caller posts the result blob it got from /api/vendor-reasons.
    result: dict


class PatchReasonCardBody(BaseModel):
    label: Optional[str] = Field(None, max_length=200)
    hidden: Optional[bool] = None


@router.get("/api/vendor-reason-cards")
async def list_reason_cards(
    include_hidden: int = 0,
    session: AsyncSession = Depends(get_session),
):
    return {
        "cards": await list_saved_cards(session, include_hidden=bool(include_hidden)),
    }


def _unwrap_reasons(stored):
    """Tolerate both shapes: old cards have `reasons` = bare list, new
    cards have `reasons` = {reasons: [...], simple_responses: {...}}.
    Returns (reasons_list, simple_responses_dict)."""
    if isinstance(stored, list):
        return stored, {"count": 0, "examples": []}
    if isinstance(stored, dict):
        return (
            stored.get("reasons") or [],
            stored.get("simple_responses") or {"count": 0, "examples": []},
        )
    return [], {"count": 0, "examples": []}


@router.get("/api/vendor-reason-cards/{card_id}")
async def get_reason_card(
    card_id: int, session: AsyncSession = Depends(get_session)
):
    card = await get_card(session, card_id)
    if not card:
        raise HTTPException(404, "card not found")
    reasons, simple = _unwrap_reasons(card.reasons)
    return {
        "id": card.id,
        "vendor_key": card.vendor_key,
        "vendor_display": card.vendor_display,
        "category_name": card.category_name,
        "band": card.band,
        "label": card.label,
        "model_used": card.model_used,
        "sample_size": card.sample_size,
        "source_ids_snapshot": card.source_ids_snapshot,
        "reasons": reasons,
        "simple_responses": simple,
        "hidden": card.hidden,
        "created_at": card.created_at.isoformat() if card.created_at else None,
        "updated_at": card.updated_at.isoformat() if card.updated_at else None,
    }


@router.post("/api/vendor-reason-cards")
async def save_reason_card(
    body: SaveReasonCardBody,
    session: AsyncSession = Depends(get_session),
):
    try:
        card = await save_card(session, result=body.result, label=body.label)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {
        "id": card.id,
        "vendor_key": card.vendor_key,
        "category_name": card.category_name,
        "band": card.band,
        "label": card.label,
    }


@router.post("/api/vendor-reason-cards/{card_id}/reanalyze")
async def reanalyze_reason_card(
    card_id: int,
    summary_lang: str = Query("en"),
    model: Optional[str] = None,
    session: AsyncSession = Depends(get_session),
):
    try:
        card = await reanalyze_card(
            session, card_id, summary_lang=summary_lang, model=model
        )
    except RuntimeError as e:
        raise HTTPException(503, str(e))
    if not card:
        raise HTTPException(404, "card not found")
    reasons, simple = _unwrap_reasons(card.reasons)
    return {
        "id": card.id,
        "vendor_key": card.vendor_key,
        "vendor_display": card.vendor_display,
        "category_name": card.category_name,
        "band": card.band,
        "label": card.label,
        "model_used": card.model_used,
        "sample_size": card.sample_size,
        "source_ids_snapshot": card.source_ids_snapshot,
        "reasons": reasons,
        "simple_responses": simple,
        "updated_at": card.updated_at.isoformat() if card.updated_at else None,
    }


@router.patch("/api/vendor-reason-cards/{card_id}")
async def patch_reason_card(
    card_id: int,
    body: PatchReasonCardBody,
    session: AsyncSession = Depends(get_session),
):
    card = None
    if body.label is not None:
        try:
            card = await update_card_label(session, card_id, body.label)
        except ValueError as e:
            raise HTTPException(400, str(e))
        if not card:
            raise HTTPException(404, "card not found")
    if body.hidden is not None:
        card = await toggle_card_hidden(session, card_id, body.hidden)
        if not card:
            raise HTTPException(404, "card not found")
    if card is None:
        raise HTTPException(400, "nothing to update")
    return {"id": card.id, "label": card.label, "hidden": card.hidden}


@router.delete("/api/vendor-reason-cards/{card_id}")
async def delete_reason_card(
    card_id: int, session: AsyncSession = Depends(get_session)
):
    ok = await delete_card(session, card_id)
    if not ok:
        raise HTTPException(404, "card not found")
    return {"deleted": card_id}


# ----------------------------------------------------------------------------
# Per-reason "show all reviews" expansion
# ----------------------------------------------------------------------------
#
# When the user clicks a reason row in the modal, the frontend POSTs the
# list of review_ids the LLM assigned to that reason and gets back the
# full review texts. No additional LLM call — the IDs were already
# captured at extract time.

import csv  # noqa: E402
import io  # noqa: E402

from fastapi.responses import Response  # noqa: E402
from sqlalchemy import select  # noqa: E402 — local import keeps the
# top-of-file imports clean while this section is self-contained.
from app.models import Analysis, Review, VendorReasonCard  # noqa: E402


@router.get("/api/vendor-reasons/reviews-by-ids")
async def reviews_by_ids(
    ids: str = Query(..., description="comma-separated review ids"),
    session: AsyncSession = Depends(get_session),
):
    """Hydrate a list of review_ids into their full text + metadata.
    Pure DB read, no LLM call. Used by the reason-row click expand
    so the user can see the entire population of reviews that fed
    a single reason.
    """
    try:
        id_list = [int(x) for x in ids.split(",") if x.strip()]
    except ValueError:
        raise HTTPException(400, "invalid ids — must be comma-separated integers")
    if not id_list:
        return {"reviews": []}
    # Safety cap. Even very loose categories rarely tag 500+ reviews to
    # a single causal reason; anything bigger is almost certainly a bug.
    if len(id_list) > 500:
        id_list = id_list[:500]

    rows = (
        await session.execute(
            select(
                Review.id,
                Review.text,
                Review.rating,
                Review.author,
                Review.posted_at,
                Review.collected_at,
                Analysis.sentiment,
                Analysis.sentiment_score,
            )
            .join(Analysis, Analysis.review_id == Review.id, isouter=True)
            .where(Review.id.in_(id_list))
        )
    ).all()

    by_id: dict[int, dict] = {}
    for rid, text, rating, author, posted, collected, sent, sscore in rows:
        s_key = sent.value if hasattr(sent, "value") else (str(sent) if sent else None)
        by_id[int(rid)] = {
            "id": int(rid),
            "text": text or "",
            "rating": int(rating) if rating is not None else None,
            "author": author or "",
            "posted_at": posted.isoformat() if posted else None,
            "collected_at": collected.isoformat() if collected else None,
            "sentiment": s_key,
            "sentiment_score": int(sscore) if sscore is not None else None,
        }
    # Preserve the request order so the user sees the same sequence the
    # LLM produced (typically rough strength-of-signal order).
    ordered = [by_id[i] for i in id_list if i in by_id]
    return {"reviews": ordered, "found": len(ordered), "requested": len(id_list)}


# ----------------------------------------------------------------------------
# CSV export
# ----------------------------------------------------------------------------


def _export_filename(ext: str) -> str:
    """Stamp the export filename with the current local date+time so
    repeated downloads land in the Downloads folder side-by-side instead
    of overwriting each other. Format: `vendor_analysis_YYYYMMDD_HHMMSS.ext`.
    ASCII-only so the Content-Disposition header is happy without RFC
    5987 encoding."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"vendor_analysis_{ts}.{ext}"


_EXPORT_HEADER = [
    "vendor",
    "type",
    "category",
    "pct",
    "count",
    "wilson_score",
    "description",
    "small_sample",
    "reasons",
]

# Fixed-column subset (everything except the trailing reasons slot).
# The CSV format keeps the trailing single "reasons" column for backward
# compat; the XLSX format replaces it with N "reason 1", "reason 2",…
# columns so each reason gets its own cell.
_EXPORT_HEADER_FIXED = _EXPORT_HEADER[:-1]


def _join_reasons(entries: list[tuple[str, int]]) -> str:
    """Pack a row's reason entries into one cell as 'text(N); text(M); …'.
    Same format the v3 parser already handles."""
    return "; ".join(f"{name}({n})" for name, n in entries)


async def _build_export_rows(
    session: AsyncSession,
    *,
    vendor_keys: str,
    include: str,
    include_reasons: int,
    exclude_weak: int,
) -> list[list]:
    """Common row-builder shared by the CSV + XLSX endpoints.

    Returns the body rows (header is added by the caller per format).
    Filters / lookups stay in one place so the two formats can never
    drift out of sync.
    """
    vendors = await list_vendors(session)
    selected = {k.strip() for k in vendor_keys.split(",") if k.strip()}
    if selected:
        vendors = [v for v in vendors if v.get("key") in selected]

    reason_map: dict[tuple[str, str, str], list[tuple[str, int]]] = {}
    if include_reasons:
        cards = (
            await session.execute(
                select(VendorReasonCard).where(VendorReasonCard.hidden.is_(False))
            )
        ).scalars().all()
        for c in cards:
            payload = c.reasons
            if isinstance(payload, dict):
                reasons = payload.get("reasons") or []
            elif isinstance(payload, list):
                reasons = payload
            else:
                reasons = []
            entries = [
                (str(r.get("reason") or "").strip(), int(r.get("count") or 0))
                for r in reasons
                if isinstance(r, dict) and (r.get("reason") or "").strip()
            ]
            key = (c.vendor_key, (c.category_name or "").strip().lower(), c.band)
            reason_map[key] = entries

    # 30 % threshold matches the gray-out logic in vendors.html, so the
    # export and the on-screen "strong" rows are the same set when the
    # user keeps the default exclude_weak=1.
    WEAK_THRESHOLD = 0.30

    # Each row's trailing cell is now a LIST of (text, count) reason
    # entries — callers join (CSV) or spread to columns (XLSX). Empty
    # list when include_reasons=0.
    rows: list[list] = []
    for v in vendors:
        if include in ("strengths", "both"):
            for s in v.get("strengths") or []:
                pos_pct = float(s.get("pos_pct") or 0.0)
                if exclude_weak and pos_pct < WEAK_THRESHOLD:
                    continue
                name = (s.get("name") or "").strip()
                key = (v.get("key"), name.lower(), "positive")
                rows.append([
                    v.get("display") or v.get("key"),
                    "strength",
                    name,
                    round(pos_pct * 100, 1),
                    int(s.get("total") or 0),
                    round(float(s.get("pos_score") or 0.0), 3),
                    (s.get("description") or "").strip(),
                    "Y" if s.get("small_sample") else "",
                    reason_map.get(key, []) if include_reasons else [],
                ])
        if include in ("weaknesses", "both"):
            for w in v.get("weaknesses") or []:
                neg_pct = float(w.get("neg_pct") or 0.0)
                if exclude_weak and neg_pct < WEAK_THRESHOLD:
                    continue
                name = (w.get("name") or "").strip()
                key = (v.get("key"), name.lower(), "negative")
                rows.append([
                    v.get("display") or v.get("key"),
                    "weakness",
                    name,
                    round(neg_pct * 100, 1),
                    int(w.get("total") or 0),
                    round(float(w.get("neg_score") or 0.0), 3),
                    (w.get("description") or "").strip(),
                    "Y" if w.get("small_sample") else "",
                    reason_map.get(key, []) if include_reasons else [],
                ])
    return rows


@router.get("/vendors/export.csv")
async def vendors_export_csv(
    vendor_keys: str = Query(
        "", description="comma-separated vendor keys; empty means all visible vendors"
    ),
    include: str = Query("both", pattern="^(strengths|weaknesses|both)$"),
    include_reasons: int = Query(
        0, description="1 = append saved reason analyses as a 'reasons' column"
    ),
    exclude_weak: int = Query(
        1,
        description=(
            "1 = drop rows whose polarity ratio is under 30 % (the same "
            "threshold the /vendors page grays out as a weak signal); 0 "
            "= keep every Top-5 entry. Default 1."
        ),
    ),
    session: AsyncSession = Depends(get_session),
):
    """Download as UTF-8 CSV (with BOM so Excel-on-Windows reads it
    cleanly). Filename is ASCII to dodge the latin-1 Content-Disposition
    pitfall."""
    body_rows = await _build_export_rows(
        session,
        vendor_keys=vendor_keys,
        include=include,
        include_reasons=include_reasons,
        exclude_weak=exclude_weak,
    )
    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(_EXPORT_HEADER)
    # CSV keeps the single-column "reasons" format for backward compat
    # (v3 parser, existing user workflows). Join the entries inline.
    for r in body_rows:
        writer.writerow(r[:-1] + [_join_reasons(r[-1])])
    # BOM lets Excel-on-Windows auto-detect UTF-8 for the Korean cells.
    body = ("﻿" + buf.getvalue()).encode("utf-8")
    return Response(
        content=body,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{_export_filename("csv")}"'
            )
        },
    )


@router.get("/vendors/export.xlsx")
async def vendors_export_xlsx(
    vendor_keys: str = Query(""),
    include: str = Query("both", pattern="^(strengths|weaknesses|both)$"),
    include_reasons: int = Query(0),
    exclude_weak: int = Query(1),
    session: AsyncSession = Depends(get_session),
):
    """Same data as the CSV endpoint, but emitted as a real .xlsx file
    via openpyxl. Excel users get formatted columns, autofilter, frozen
    header, and proper number/text typing instead of "everything is
    text and the leading-zero categories look broken." """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    body_rows = await _build_export_rows(
        session,
        vendor_keys=vendor_keys,
        include=include,
        include_reasons=include_reasons,
        exclude_weak=exclude_weak,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "vendor_analysis"

    # XLSX-specific layout: one ROW per reason. A category with N
    # reasons becomes N rows; the fixed cells (vendor / type /
    # category / pct / count / wilson / description / small_sample)
    # repeat on every row so Excel's autofilter & sort still work
    # cleanly. The trailing pair is "reason" (text(N) form, kept so the
    # v3 parser's existing regex still works) followed by a NUMERIC
    # "review_count" cell — Excel users asked for the integer count in
    # its own cell so they can sort / pivot / chart on it directly
    # without first regex-stripping the "(N)" suffix out of the reason
    # text. The redundancy is intentional: v3 parser ignores
    # review_count, Excel users ignore the "(N)" suffix.
    full_header = _EXPORT_HEADER_FIXED + ["reason", "review_count"]
    fixed_len = len(_EXPORT_HEADER_FIXED)

    # Header row — bold + light gray fill so the autofilter dropdown
    # arrows show up against a backdrop.
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    for col_idx, label in enumerate(full_header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Body — expand each category row into per-reason rows. Categories
    # with zero saved reasons still emit one row (blank reason cell)
    # so the strength/weakness header itself doesn't disappear.
    ws_row = 2
    for row in body_rows:
        fixed_cells = row[:fixed_len]
        entries = row[-1] or []
        if not entries:
            for c_idx, value in enumerate(fixed_cells, start=1):
                ws.cell(row=ws_row, column=c_idx, value=value)
            ws_row += 1
        else:
            for name, n in entries:
                for c_idx, value in enumerate(fixed_cells, start=1):
                    ws.cell(row=ws_row, column=c_idx, value=value)
                ws.cell(row=ws_row, column=fixed_len + 1, value=f"{name}({n})")
                # Numeric typing so Excel treats it as a real number
                # (SUM / AVERAGE / sort all work). int(n) cast is
                # defensive — n is already int from _build_export_rows
                # but the source may evolve.
                ws.cell(row=ws_row, column=fixed_len + 2, value=int(n))
                ws_row += 1

    # Column widths tuned for typical Korean content.
    widths = {
        "vendor": 28, "type": 10, "category": 32,
        "pct": 8, "count": 8, "wilson_score": 12,
        "description": 40, "small_sample": 12,
    }
    for col_idx, label in enumerate(_EXPORT_HEADER_FIXED, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(label, 16)
    ws.column_dimensions[get_column_letter(fixed_len + 1)].width = 60   # reason
    ws.column_dimensions[get_column_letter(fixed_len + 2)].width = 14   # review_count

    # Quality-of-life: freeze the header row, enable autofilter on the
    # whole table, left-align everything for Korean readability.
    ws.freeze_panes = "A2"
    last_col_letter = get_column_letter(len(full_header))
    last_row = max(1, ws_row - 1)
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    buf = io.BytesIO()
    wb.save(buf)
    body = buf.getvalue()
    return Response(
        content=body,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                f'attachment; filename="{_export_filename("xlsx")}"'
            )
        },
    )
