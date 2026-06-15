"""Competitive-v3 — read-only "what did each vendor's users say?" view.

The user uploads a CSV or Excel file exported from /vendors and the
page shows, per vendor, every strength category + its `reasons` list
(causal mechanisms with counts). Clicking a reason opens a popup with
the actual review snippets that fed that reason in the original
/vendors per-strength analysis.

No DB persistence for v3 itself — the page is a viewer. The popup
lookup chases the VendorReasonCard that produced the row's reasons
(matched by vendor_key + lowercased category_name + band='positive')
and uses its stored review_ids to fetch real review text.

CSV and Excel uploads both supported; the parser dispatches by file
extension/content-type.
"""
from __future__ import annotations

import csv as _csv
import hashlib
import io
import json
import re
import time
from typing import Any, Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.models import Analysis, CompetitiveV3Card, Review, VendorReasonCard
from app.services.vendors import list_vendors, _vendor_key


# Same convention used by the v2 reasons parser — "reason text (N)".
_REASON_COUNT_RE = re.compile(r"^(.+?)\s*\((\d+)\)\s*$")


def parse_reasons_cell(cell: str) -> list[dict]:
    """Split a reasons cell ("이유A(12); 이유B(8); …") into structured
    entries. Returns [{text, count}, ...]."""
    out: list[dict] = []
    for piece in (cell or "").split(";"):
        p = piece.strip()
        if not p:
            continue
        m = _REASON_COUNT_RE.match(p)
        if m:
            text = m.group(1).strip()
            try:
                count = int(m.group(2))
            except ValueError:
                count = 1
        else:
            text = p
            count = 1
        if text:
            out.append({"text": text[:300], "count": max(0, count)})
    return out


def _merge_split_reason_rows(rows: list[dict]) -> list[dict]:
    """Fold rows from the row-per-reason XLSX layout back into a single
    row per (vendor, type, category).

    The new XLSX format emits one row per reason — vendor/category/pct
    cells repeat, the trailing column is a singular `reason` cell with
    one entry. This function detects that layout (presence of `reason`
    or absence of `reasons`) and concatenates the entries into a single
    `reasons` cell joined by "; ", so build_view sees the canonical
    shape regardless of which export format the user uploaded.

    Legacy CSV / one-row-per-category XLSX uploads pass through
    unchanged because each (vendor, type, category) key appears only
    once and already has its `reasons` cell populated.
    """
    merged: dict[tuple[str, str, str], dict] = {}
    order: list[tuple[str, str, str]] = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        key = (
            (r.get("vendor") or "").strip(),
            (r.get("type") or "").strip(),
            (r.get("category") or "").strip(),
        )
        if not any(key):
            # Junk row — let downstream skip it
            order.append(key)
            merged[key] = r
            continue
        single = (r.get("reason") or "").strip()
        if key not in merged:
            merged[key] = dict(r)
            order.append(key)
            if "reasons" not in merged[key]:
                merged[key]["reasons"] = ""
        if single:
            existing = (merged[key].get("reasons") or "").strip()
            merged[key]["reasons"] = (
                f"{existing}; {single}" if existing else single
            )
    return [merged[k] for k in order]


def _score_sheet(headers: list[str]) -> int:
    """Score how likely a sheet is the 'vendor_analysis' export.
    Higher = better match. Used to pick the right sheet when an
    uploaded workbook has multiple sheets (e.g. summary +
    vendor_analysis from the user's categorized file)."""
    low = [str(h or "").strip().lower() for h in headers]
    raw = [str(h or "").strip() for h in headers]
    score = 0
    if "vendor" in low: score += 3
    if "category" in low: score += 3
    if "type" in low: score += 2
    if "reason" in low or "reasons" in low: score += 2
    if "count" in low: score += 1
    # The categorized-mode 카테고리 column is a strong signal that this
    # is the right sheet (vs. the summary pivot, which uses other
    # Korean labels but rarely the exact "카테고리" string as a header).
    if "카테고리" in raw and "vendor" in low: score += 2
    return score


def _pick_main_sheet(wb):
    """Choose the sheet that looks like vendor_analysis. openpyxl's
    `wb.active` reflects whichever sheet was selected at save time,
    which for the user's file is 'summary' — not what we want. Scan
    every sheet, score by header keywords, pick the highest."""
    best = None
    best_score = -1
    for name in wb.sheetnames:
        ws = wb[name]
        headers_row = next(ws.iter_rows(values_only=True, max_row=1), None) or ()
        score = _score_sheet(list(headers_row))
        if score > best_score:
            best_score = score
            best = ws
    return best if best else wb.active


def parse_uploaded_file(filename: str, content: bytes) -> list[dict]:
    """Read a /vendors-export-shaped CSV or XLSX into a list of dict
    rows. Dispatch by extension; both shapes have the same column names
    so downstream code doesn't care which path produced the data.

    Returns RAW per-reason rows (no row-merge). Caller decides:
      - build_view() merges per (vendor, type, category) for the
        per-vendor display.
      - build_categorized_view() keeps rows separate and groups by
        the top-level 카테고리 column instead.

    Returns rows as plain dicts with string values (caller can coerce).
    """
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = _pick_main_sheet(wb)
        rows_iter = ws.iter_rows(values_only=True)
        headers_row = next(rows_iter, None)
        if not headers_row:
            return []
        headers = [str(h or "").strip() for h in headers_row]
        out: list[dict] = []
        for r in rows_iter:
            if not r:
                continue
            entry = {}
            for i, h in enumerate(headers):
                if i >= len(r):
                    continue
                v = r[i]
                entry[h] = "" if v is None else str(v)
            if entry:
                out.append(entry)
        return out
    # default: CSV (handles BOM via utf-8-sig)
    text = content.decode("utf-8-sig", errors="replace")
    reader = _csv.DictReader(io.StringIO(text))
    return [dict(r) for r in reader]


def has_top_category(rows: list[dict]) -> bool:
    """True when any uploaded row carries a non-empty 카테고리 column —
    the signal that the file came from the categorized-export workflow
    and should be rendered with the per-category aggregation view."""
    for r in rows or []:
        if isinstance(r, dict) and (r.get("카테고리") or "").strip():
            return True
    return False


def _row_band_from_type(t: str) -> str:
    t = (t or "").strip().lower()
    if t == "weakness":
        return "negative"
    return "positive"


async def build_view(
    session: AsyncSession,
    raw_rows: list[dict],
) -> dict:
    """Turn parsed rows into the structure the template renders.

    Accepts raw per-reason rows; calls _merge_split_reason_rows first
    so it sees the canonical one-row-per-(vendor,type,category) shape
    with a semicolon-joined `reasons` cell.

    Pipeline:
      1. Merge per-reason rows back into per-category rows.
      2. Normalize each row, filter to those with reasons.
      3. Look up vendor logo + canonical key by matching the row's
         vendor display string against list_vendors() output.
      4. Group rows by canonical vendor key, then by (category, band)
         within that vendor. Each group keeps its parsed reasons +
         the originating row metadata for the popup lookup later.

    Returns:
      {
        "input_row_count":     int,
        "rows_with_reasons":   int,
        "vendors_with_reasons": int,
        "skipped_no_reasons":  int,
        "vendors": [
          {
            "key":          str,    # canonical vendor key
            "display":      str,    # display name
            "icon_url":     str|None,
            "categories": [
              {
                "category": str,
                "band":     "positive"|"negative",
                "pct":      float,
                "count":    int,
                "reasons":  [{"text", "count"}, ...],
              }, ...
            ]
          }, ...
        ]
      }
    """
    # Build a display-name → (canonical_key, icon_url) lookup so we can
    # enrich CSV rows with logos. Match case-insensitively because the
    # CSV display string came from the same list_vendors() output that
    # we're now reading back here.
    vendors_db = await list_vendors(session)
    by_display: dict[str, dict] = {}
    by_key: dict[str, dict] = {}
    for v in vendors_db:
        rec = {"key": v.get("key"), "icon_url": v.get("icon_url"), "display": v.get("display")}
        if v.get("display"):
            by_display[str(v["display"]).strip().lower()] = rec
        if v.get("key"):
            by_key[str(v["key"]).strip().lower()] = rec

    # parse_uploaded_file now returns RAW rows so build_categorized_view
    # can group by reason. The per-vendor view here still wants merged
    # rows (one per vendor,type,category with joined `reasons`), so
    # collapse here.
    raw_rows = _merge_split_reason_rows(raw_rows or [])

    grouped: dict[str, dict] = {}
    rows_with_reasons = 0
    skipped_no_reasons = 0
    input_row_count = 0

    for raw in raw_rows or []:
        if not isinstance(raw, dict):
            continue
        input_row_count += 1
        vendor_display = str(raw.get("vendor") or "").strip()
        category = str(raw.get("category") or "").strip()
        if not vendor_display or not category:
            continue
        reasons_cell = str(raw.get("reasons") or "").strip()
        reasons = parse_reasons_cell(reasons_cell)
        if not reasons:
            skipped_no_reasons += 1
            continue
        rows_with_reasons += 1
        # Resolve canonical vendor key + logo. Try display match first
        # (covers exact /vendors export rows), then fall back to a key
        # match in case the CSV came from a workspace that renamed
        # things.
        rec = by_display.get(vendor_display.lower()) or by_key.get(vendor_display.lower())
        if rec:
            vendor_key = rec["key"]
            icon_url = rec.get("icon_url")
            display = rec.get("display") or vendor_display
        else:
            # Fallback: derive a stable-ish key from the display name
            # using the same stemmer the live page uses; logos remain
            # absent but the user still sees the data.
            vendor_key = _vendor_key(vendor_display, vendor_display)
            icon_url = None
            display = vendor_display

        try:
            pct = float(raw.get("pct") or 0)
        except (TypeError, ValueError):
            pct = 0.0
        try:
            count = int(raw.get("count") or 0)
        except (TypeError, ValueError):
            count = 0
        band = _row_band_from_type(str(raw.get("type") or ""))

        vendor_node = grouped.setdefault(vendor_key, {
            "key": vendor_key,
            "display": display,
            "icon_url": icon_url,
            "categories": [],
        })
        # Prefer the longest display we encounter (mirrors list_vendors
        # behavior so users see "Weight Watchers Program" rather than a
        # store-stub).
        if len(display) > len(vendor_node["display"] or ""):
            vendor_node["display"] = display
        if icon_url and not vendor_node["icon_url"]:
            vendor_node["icon_url"] = icon_url
        vendor_node["categories"].append({
            "category": category,
            "band": band,
            "pct": pct,
            "count": count,
            "reasons": reasons,
        })

    # Stable display order: vendors by display name; within a vendor,
    # categories by band (positive first) then by pct desc.
    vendors_out = sorted(
        grouped.values(),
        key=lambda v: (v["display"] or "").lower(),
    )
    for v in vendors_out:
        v["categories"].sort(
            key=lambda c: (
                0 if c["band"] == "positive" else 1,
                -float(c["pct"] or 0),
            )
        )

    return {
        "input_row_count": input_row_count,
        "rows_with_reasons": rows_with_reasons,
        "vendors_with_reasons": len(vendors_out),
        "skipped_no_reasons": skipped_no_reasons,
        "vendors": vendors_out,
    }


# ----------------------------------------------------------------------------
# LLM categorization — assign each reason to one of ~10 cross-vendor categories
# ----------------------------------------------------------------------------

# In-memory cache: hash of (reason text list) → categorized rows. Lets a
# user re-upload the same export without burning tokens. 30-min TTL so
# Render redeploys naturally clear it.
_CAT_CACHE_TTL_SECONDS = 30 * 60
_categorize_cache: dict[str, tuple[float, list[dict]]] = {}

# Short-lived buffer of "current upload" results so the Save / Export
# buttons in the result partial can recover the full rows + computed
# view without re-uploading. Keyed by the same input hash returned in
# the result payload. Trimmed naturally by TTL (also 30 min).
_RECENT_UPLOADS_TTL = 30 * 60
_recent_uploads: dict[str, tuple[float, dict]] = {}  # hash → {rows, result, filename}


def remember_upload(input_hash: str, rows: list[dict], result: dict, filename: str) -> None:
    _recent_uploads[input_hash] = (
        time.time(),
        {"rows": rows, "result": result, "filename": filename},
    )


def recall_upload(input_hash: str) -> Optional[dict]:
    entry = _recent_uploads.get(input_hash)
    if not entry:
        return None
    ts, data = entry
    if time.time() - ts > _RECENT_UPLOADS_TTL:
        _recent_uploads.pop(input_hash, None)
        return None
    return data


def hash_rows(rows: list[dict]) -> str:
    """Stable SHA256 of the canonical (reason, vendor, type, category, count)
    set for an upload. Used both as the categorize-cache key and as the
    short-lived 'recent uploads' lookup key returned to the template."""
    canon: list[tuple] = []
    for r in rows or []:
        if not isinstance(r, dict):
            continue
        canon.append((
            str(r.get("reason") or "").strip(),
            str(r.get("vendor") or "").strip(),
            str(r.get("type") or "").strip(),
            str(r.get("category") or "").strip(),
            str(r.get("review_count") or r.get("count") or "0"),
        ))
    canon.sort()
    return hashlib.sha256(
        json.dumps(canon, ensure_ascii=False).encode("utf-8")
    ).hexdigest()


def _strip_fences(s: str) -> str:
    s = (s or "").strip()
    if s.startswith("```"):
        s = re.sub(r"^```(?:json)?\s*", "", s)
        s = re.sub(r"\s*```\s*$", "", s)
    return s.strip()


# Bump this whenever _build_categorize_prompt changes — its value goes
# into the cache key so cached results from the prior prompt version are
# automatically invalidated. Saves users from "I changed the rules but
# nothing happened" confusion.
_CATEGORIZE_PROMPT_VERSION = "v3-2026-06-feature-specific-usability-routing"


def _categorize_cache_key(items: list[dict], lang: str, model: str) -> str:
    """Hash of the canonical (reason text + vendor + count) tuples plus
    lang + model + prompt version — two uploads of the same data with
    the same prompt should hit the cache; prompt changes invalidate."""
    canon = json.dumps(
        sorted(
            (i["reason"], i["vendor"], int(i["count"] or 0)) for i in items
        ),
        ensure_ascii=False,
    )
    return hashlib.sha256(
        f"{_CATEGORIZE_PROMPT_VERSION}|{model}|{lang}|{canon}".encode("utf-8")
    ).hexdigest()


def _build_categorize_prompt(
    items: list[dict],
    lang: str,
    target_categories: int = 10,
) -> tuple[str, str]:
    """Build the (system, user) prompt for the categorize-reasons call.

    items: list of {i, reason, vendor, vendor_category, count}
    Returns (system_prompt, user_message).
    """
    lang_label = {"ko": "Korean", "en": "English"}.get(lang, "Korean")
    total_reviews = sum(int(i.get("count") or 0) for i in items)
    vendor_set = sorted({i["vendor"] for i in items if i.get("vendor")})

    system = (
        f"You are categorizing user-review REASONS (causal mechanisms for\n"
        f"product strengths) into a small set of cross-vendor top-level\n"
        f"categories. The dataset comes from {len(vendor_set)} vendors with\n"
        f"{len(items)} reasons covering {total_reviews:,} reviews.\n\n"
        f"Vendors: {', '.join(vendor_set)}\n\n"
        f"GOAL\n"
        f"  1. Discover ~{target_categories} top-level categories that group\n"
        f"     these reasons by THEME (what aspect of the product the reason\n"
        f"     describes). Aim for {max(8, target_categories - 2)}–"
        f"{target_categories + 2} categories — go fewer when reasons cluster\n"
        f"     tightly, more only when needed to avoid heterogeneous buckets.\n"
        f"  2. Categories must be CROSS-VENDOR — every category should contain\n"
        f"     reasons from at least 2 different vendors. Avoid single-vendor\n"
        f"     buckets unless the theme is genuinely vendor-specific.\n"
        f"  3. Assign EVERY input reason to exactly one category.\n\n"
        f"CATEGORY NAMING RULES\n"
        f"  - 4–18 character noun phrases in {lang_label}\n"
        f"  - Use middle dot '·' to join related concepts (e.g. '칼로리·매크로·"
        f"영양소 추적')\n"
        f"  - VALUE-NEUTRAL — describe the theme, not whether it's positive\n"
        f"    or negative. NOT '훌륭한 X', NOT 'X 문제'\n"
        f"  - Reference examples from the user's prior dataset for tone:\n"
        f"      '칼로리·매크로·영양소 추적'\n"
        f"      '사진·바코드·음성 음식 입력'\n"
        f"      '동기부여·책임감·습관 형성'\n"
        f"      '운동·수면·건강 데이터 추적'\n"
        f"      'UI/UX·사용 편의성'\n"
        f"      '음식 DB·검색·레시피'\n"
        f"      '교육·레슨·강사 콘텐츠'\n"
        f"      '게임화·캐릭터·보상'\n"
        f"      '무료·가격 가치'\n"
        f"      '개인화 코칭·인사이트'\n"
        f"      '커뮤니티·소셜'\n"
        f"    Re-use these names when a new reason cleanly fits; coin new\n"
        f"    ones only when none of the above match.\n\n"
        f"SPECIAL MERGE RULES (apply BEFORE coining new categories):\n"
        f"  - Wearable / device DATA collection (Apple Watch, Fitbit,\n"
        f"    WHOOP, Oura ring → heart rate / steps / sleep / HRV /\n"
        f"    workout metrics) → MERGE INTO '운동·수면·건강 데이터 추적'.\n"
        f"    Don't split this off as its own 'device integration'\n"
        f"    bucket — the value the user gets is the health data, not\n"
        f"    the act of pairing.\n"
        f"  - APP integration / interop (Apple Health / Google Fit /\n"
        f"    MyFitnessPal / Strava sync, data export, ecosystem fit)\n"
        f"    → SEPARATE category. The value is data portability and\n"
        f"    workflow continuity across apps, not health data per se.\n"
        f"    Suggested name: '외부 앱 연동·데이터 호환'. Create it only\n"
        f"    when ≥2 vendors have reasons fitting this distinct value.\n\n"
        f"FEATURE-SPECIFIC USABILITY RULE — IMPORTANT:\n"
        f"  When a reason describes how EASY / CONVENIENT / INTUITIVE\n"
        f"  a SPECIFIC FEATURE is (the noun phrase identifies WHAT is\n"
        f"  easy, not just THAT it's easy), route to that feature's\n"
        f"  category — NOT to 'UI/UX·사용 편의성'. The UX bucket should\n"
        f"  only collect reasons about APP-WIDE usability with no\n"
        f"  specific feature attached.\n"
        f"  Examples — route these to the FEATURE category:\n"
        f"    ✗ NOT UI/UX:   '사진 음식 인식의 편리함'           → 사진·바코드·음성 음식 입력\n"
        f"    ✗ NOT UI/UX:   '바코드 스캐너로 빠른 입력'         → 사진·바코드·음성 음식 입력\n"
        f"    ✗ NOT UI/UX:   '음성 기록의 편의성'                → 사진·바코드·음성 음식 입력\n"
        f"    ✗ NOT UI/UX:   '음식 검색이 빠름'                  → 음식 DB·검색·레시피\n"
        f"    ✗ NOT UI/UX:   '레시피 추가가 간편'                → 음식 DB·검색·레시피\n"
        f"    ✗ NOT UI/UX:   '운동 기록이 자동으로 동기화돼 편리' → 운동·수면·건강 데이터 추적\n"
        f"    ✗ NOT UI/UX:   '칼로리 자동 계산이 편리'           → 칼로리·매크로·영양소 추적\n"
        f"    ✗ NOT UI/UX:   '미팅 참석이 동기부여에 효과적'     → 동기부여·책임감·습관 형성\n"
        f"    ✗ NOT UI/UX:   '커뮤니티 참여가 쉬움'              → 커뮤니티·소셜\n"
        f"  Examples — these BELONG in UI/UX·사용 편의성:\n"
        f"    ✓ '직관적 메뉴 구조와 깔끔한 디자인'\n"
        f"    ✓ '앱 전반의 빠른 응답 속도'\n"
        f"    ✓ '대시보드 한눈에 보이는 정보 배치'\n"
        f"    ✓ '다크 모드 / 색상 테마 선택'\n"
        f"    ✓ '광고 없는 깔끔한 화면'\n"
        f"  Heuristic: if you can replace 'X 편리함' with the FEATURE\n"
        f"  category's working name and the sentence still makes sense,\n"
        f"  route there. UI/UX is the LAST resort — pick a feature\n"
        f"  category first whenever a specific feature is named.\n\n"
        f"OUTPUT — JSON only, no prose, no markdown fences:\n"
        f"{{\n"
        f'  "categories": ["category A", "category B", ...],\n'
        f'  "assignments": [\n'
        f'    {{"i": 1, "category": "category A"}},\n'
        f'    {{"i": 2, "category": "category B"}},\n'
        f"    ...\n"
        f"  ]\n"
        f"}}\n\n"
        f"`assignments` length MUST equal the number of input reasons\n"
        f"({len(items)}). Every `category` value MUST appear in `categories`.\n"
        f"Indexes MUST be 1..{len(items)} — no skips, no duplicates."
    )

    lines = []
    for it in items:
        lines.append(
            f"[{it['i']}] {it['reason']}"
            f" | vendor={it['vendor']}"
            f" | from={it['vendor_category']}"
            f" | reviews={it['count']}"
        )
    user_msg = "Reasons to categorize:\n" + "\n".join(lines)
    return system, user_msg


async def categorize_reasons_with_llm(
    rows: list[dict],
    *,
    lang: str = "ko",
    model: Optional[str] = None,
    target_categories: int = 10,
) -> list[dict]:
    """Run a Claude call that classifies every reason into one of
    ~target_categories cross-vendor top-level categories. Mutates each
    row in place to add a "카테고리" key. Returns the same list.

    The function expects RAW per-reason rows (one row per reason),
    which is what parse_uploaded_file returns. Rows missing a usable
    reason cell are skipped.
    """
    if not settings.ANTHROPIC_API_KEY:
        raise RuntimeError("ANTHROPIC_API_KEY is not set")

    # Extract the per-reason items the LLM will see.
    items: list[dict] = []
    item_to_row: dict[int, dict] = {}
    for raw in rows or []:
        if not isinstance(raw, dict):
            continue
        reason_text = str(raw.get("reason") or "").strip()
        if not reason_text:
            continue
        # Strip "(N)" suffix if present so the LLM sees clean text.
        m = _REASON_COUNT_RE.match(reason_text)
        if m:
            reason_text = m.group(1).strip()
        try:
            count = int(float(raw.get("review_count") or 0))
        except (TypeError, ValueError):
            count = 0
        if not count and m:
            try:
                count = int(m.group(2))
            except ValueError:
                count = 0
        vendor = str(raw.get("vendor") or "").strip()
        vendor_category = str(raw.get("category") or "").strip()
        idx = len(items) + 1
        items.append({
            "i": idx,
            "reason": reason_text,
            "vendor": vendor,
            "vendor_category": vendor_category,
            "count": count,
        })
        item_to_row[idx] = raw

    if not items:
        return rows

    chosen_model = (model or settings.ANTHROPIC_MODEL).strip()
    if chosen_model not in settings.allowed_models:
        chosen_model = settings.ANTHROPIC_MODEL

    # Cache check.
    cache_key = _categorize_cache_key(items, lang, chosen_model)
    cached = _categorize_cache.get(cache_key)
    if cached and time.time() - cached[0] < _CAT_CACHE_TTL_SECONDS:
        cat_by_idx = cached[1]
        for i, cat in cat_by_idx.items():
            if i in item_to_row:
                item_to_row[i]["카테고리"] = cat
                item_to_row[i]["_v3_cached"] = True
        return rows

    system, user_msg = _build_categorize_prompt(items, lang, target_categories)
    from anthropic import AsyncAnthropic
    client = AsyncAnthropic(api_key=settings.ANTHROPIC_API_KEY)
    resp = await client.messages.create(
        model=chosen_model,
        max_tokens=8192,
        temperature=0.0,
        system=system,
        messages=[{"role": "user", "content": user_msg}],
    )
    text = _strip_fences("".join(getattr(b, "text", "") for b in resp.content))
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError as e:
        raise RuntimeError(
            f"categorize LLM returned invalid JSON: {e}; raw={text[:200]!r}"
        )

    declared_categories = parsed.get("categories") or []
    valid_cat_set = {str(c).strip() for c in declared_categories if str(c).strip()}
    assignments_raw = parsed.get("assignments") or []

    cat_by_idx: dict[int, str] = {}
    for a in assignments_raw:
        if not isinstance(a, dict):
            continue
        try:
            i = int(a.get("i"))
        except (TypeError, ValueError):
            continue
        cat = str(a.get("category") or "").strip()
        if not cat:
            continue
        # Allow categories the LLM didn't declare in `categories` (defensive).
        if cat not in valid_cat_set:
            valid_cat_set.add(cat)
        cat_by_idx[i] = cat

    # Fallback: any reasons the LLM forgot get a "기타" bucket.
    for i in item_to_row.keys():
        if i not in cat_by_idx:
            cat_by_idx[i] = "기타"

    # Cache + apply.
    _categorize_cache[cache_key] = (time.time(), cat_by_idx)
    for i, cat in cat_by_idx.items():
        if i in item_to_row:
            item_to_row[i]["카테고리"] = cat

    return rows


# ----------------------------------------------------------------------------
# Categorized view — group reasons by top-level 카테고리 column
# ----------------------------------------------------------------------------


async def build_categorized_view(
    session: AsyncSession,
    raw_rows: list[dict],
) -> dict:
    """When the uploaded XLSX has a per-reason 카테고리 column (top-level
    grouping decided by an upstream classifier), pivot the data around
    those top categories instead of around vendors.

    Each input row is one reason. We aggregate:
      - per top category: total review_count + reason_count + list of
        vendors contributing reasons to this category
      - vendor × category matrix: how many reviews each vendor
        contributed to each top category (matches the Summary sheet's
        pivot table)

    Returns:
      {
        "categorized":   True,                # disambiguates from build_view
        "totals":        {reviews, reasons, vendors, categories},
        "categories": [
          {
            "name":          str,
            "review_count":  int,             # sum across all reasons
            "pct":           float,           # of total reviews
            "reason_count":  int,
            "vendors": [
              {key, display, icon_url, review_count, reasons: [
                {text, count, vendor_category, vendor_pct, band}
              ]}, ...
            ]
          }, ...
        ],
        "matrix": {
          "categories":   [str, ...],
          "vendors":      [{key, display, icon_url}, ...],
          "cells":        [[int, ...], ...],  # row = category, col = vendor
          "row_totals":   [int, ...],
          "col_totals":   [int, ...],
          "grand_total":  int,
        }
      }
    """
    # Vendor logo / canonical key lookup — same approach as build_view.
    vendors_db = await list_vendors(session)
    by_display: dict[str, dict] = {}
    by_key: dict[str, dict] = {}
    for v in vendors_db:
        rec = {
            "key": v.get("key"),
            "icon_url": v.get("icon_url"),
            "display": v.get("display"),
        }
        if v.get("display"):
            by_display[str(v["display"]).strip().lower()] = rec
        if v.get("key"):
            by_key[str(v["key"]).strip().lower()] = rec

    # Working dictionaries — we'll sort + materialize at the end.
    by_top: dict[str, dict] = {}
    matrix: dict[tuple[str, str], int] = {}      # (top_cat_name, vendor_key) → reviews
    vendor_meta: dict[str, dict] = {}            # vendor_key → {key, display, icon_url}
    total_reviews = 0
    reason_rows = 0

    for raw in raw_rows or []:
        if not isinstance(raw, dict):
            continue
        top = (raw.get("카테고리") or "").strip()
        if not top:
            continue
        vendor_display_raw = str(raw.get("vendor") or "").strip()
        vendor_category = str(raw.get("category") or "").strip()
        reason_text = str(raw.get("reason") or "").strip()
        if not vendor_display_raw or not vendor_category or not reason_text:
            continue

        # Per-reason review count: prefer the explicit numeric column;
        # fall back to parsing the "(N)" suffix on the reason cell when
        # the numeric column is missing or 0 (legacy CSVs).
        try:
            review_count = int(float(raw.get("review_count") or 0))
        except (TypeError, ValueError):
            review_count = 0
        m = _REASON_COUNT_RE.match(reason_text)
        if m:
            if not review_count:
                try:
                    review_count = int(m.group(2))
                except ValueError:
                    pass
            reason_text = m.group(1).strip()

        # Resolve canonical vendor + logo.
        rec = (
            by_display.get(vendor_display_raw.lower())
            or by_key.get(vendor_display_raw.lower())
        )
        if rec:
            vendor_key = rec["key"]
            icon_url = rec.get("icon_url")
            display = rec.get("display") or vendor_display_raw
        else:
            vendor_key = _vendor_key(vendor_display_raw, vendor_display_raw)
            icon_url = None
            display = vendor_display_raw

        band = (
            "negative"
            if str(raw.get("type") or "").strip().lower() == "weakness"
            else "positive"
        )
        try:
            vendor_pct = float(raw.get("pct") or 0)
        except (TypeError, ValueError):
            vendor_pct = 0.0

        # Build the nested structure.
        top_node = by_top.setdefault(top, {
            "name": top,
            "vendors": {},
            "review_count": 0,
            "reason_count": 0,
        })
        v_node = top_node["vendors"].setdefault(vendor_key, {
            "key": vendor_key,
            "display": display,
            "icon_url": icon_url,
            "review_count": 0,
            "reasons": [],
        })
        v_node["reasons"].append({
            "text": reason_text,
            "count": review_count,
            "vendor_category": vendor_category,
            "vendor_pct": vendor_pct,
            "band": band,
        })
        v_node["review_count"] += review_count
        # Prefer the longest display we've seen for this vendor (mirrors
        # build_view behaviour for cross-store mergers).
        if len(display) > len(v_node["display"] or ""):
            v_node["display"] = display
        if icon_url and not v_node.get("icon_url"):
            v_node["icon_url"] = icon_url

        top_node["review_count"] += review_count
        top_node["reason_count"] += 1
        total_reviews += review_count
        reason_rows += 1

        matrix[(top, vendor_key)] = matrix.get((top, vendor_key), 0) + review_count
        prev = vendor_meta.get(vendor_key)
        if not prev or len(display) > len(prev.get("display") or ""):
            vendor_meta[vendor_key] = {
                "key": vendor_key,
                "display": display,
                "icon_url": icon_url or (prev or {}).get("icon_url"),
            }

    # Sort: top categories by review_count desc; within each, vendors
    # by review_count desc; within each vendor, reasons by count desc.
    categories_sorted: list[dict] = []
    for top_node in sorted(by_top.values(), key=lambda c: -c["review_count"]):
        top_node["pct"] = (
            (top_node["review_count"] / total_reviews * 100)
            if total_reviews else 0.0
        )
        vendors_sorted = sorted(
            top_node["vendors"].values(),
            key=lambda v: -v["review_count"],
        )
        for v in vendors_sorted:
            v["reasons"].sort(key=lambda r: -r["count"])
        top_node["vendors"] = vendors_sorted
        categories_sorted.append(top_node)

    # Matrix: vendors sorted by total contribution desc so the most
    # active vendors land near the left edge.
    vendor_totals = {
        vk: sum(matrix.get((top["name"], vk), 0) for top in categories_sorted)
        for vk in vendor_meta
    }
    sorted_vendor_keys = sorted(
        vendor_meta.keys(),
        key=lambda vk: -vendor_totals.get(vk, 0),
    )
    sorted_top_names = [c["name"] for c in categories_sorted]

    cells: list[list[int]] = []
    for top in sorted_top_names:
        cells.append([matrix.get((top, vk), 0) for vk in sorted_vendor_keys])
    row_totals = [sum(row) for row in cells]
    col_totals = [
        sum(cells[r][c] for r in range(len(cells)))
        for c in range(len(sorted_vendor_keys))
    ]
    grand_total = sum(row_totals)
    # Precompute max cell value for the heatmap intensity — Jinja's
    # nested-loop {% set %} can't reach outer scope so this can't be
    # done in the template.
    max_cell = 0
    for row in cells:
        for n in row:
            if n > max_cell:
                max_cell = n
    if max_cell == 0:
        max_cell = 1

    return {
        "categorized": True,
        "totals": {
            "reviews": total_reviews,
            "reasons": reason_rows,
            "vendors": len(vendor_meta),
            "categories": len(categories_sorted),
        },
        "categories": categories_sorted,
        "matrix": {
            "categories": sorted_top_names,
            "vendors": [vendor_meta[vk] for vk in sorted_vendor_keys],
            "cells": cells,
            "row_totals": row_totals,
            "col_totals": col_totals,
            "grand_total": grand_total,
            "max_cell": max_cell,
        },
    }


# ----------------------------------------------------------------------------
# Competitive criteria — group the AI categories into higher-level buckets
# ----------------------------------------------------------------------------

# Bump when _build_criteria_prompt changes — invalidates cached groupings.
_CRITERIA_PROMPT_VERSION = "v3-criteria-2026-06-initial"
_CRITERIA_CACHE_TTL_SECONDS = 30 * 60
_criteria_cache: dict[str, tuple[float, dict]] = {}

# Catch-all bucket for categories the LLM forgot to assign.
_UNASSIGNED_NAME = "미분류"


def _criteria_cache_key(category_names: list[str], lang: str, model: str) -> str:
    canon = json.dumps(sorted(category_names), ensure_ascii=False)
    return hashlib.sha256(
        f"{_CRITERIA_PROMPT_VERSION}|{model}|{lang}|{canon}".encode("utf-8")
    ).hexdigest()


def _normalize_criteria(parsed: dict, category_names: list[str]) -> dict:
    """Turn a raw LLM grouping response into the canonical
    {"groups": [{"name", "categories": [...]}, ...]} shape.

    Guarantees:
      - every input category appears exactly once across all groups
      - categories the LLM dropped / mislabelled land in a trailing
        '미분류' group
      - declared-but-empty criteria are omitted
      - group order follows the LLM's declared `criteria` order, with
        any extra criteria (only seen in assignments) appended after.

    Pure function — no LLM, no I/O — so it's unit-testable.
    """
    valid = list(dict.fromkeys(n for n in category_names if n))  # de-dupe, keep order
    valid_set = set(valid)

    declared = []
    for c in (parsed.get("criteria") or []):
        name = str(c).strip()
        if name and name not in declared:
            declared.append(name)

    # category → criterion, first assignment wins.
    assigned: dict[str, str] = {}
    order_extra: list[str] = []
    for a in (parsed.get("assignments") or []):
        if not isinstance(a, dict):
            continue
        cat = str(a.get("category") or "").strip()
        crit = str(a.get("criterion") or "").strip()
        if not cat or not crit or cat not in valid_set or cat in assigned:
            continue
        assigned[cat] = crit
        if crit not in declared and crit not in order_extra:
            order_extra.append(crit)

    # Build groups in stable order: declared criteria first, then extras.
    groups: dict[str, list[str]] = {}
    for cat in valid:  # iterate categories in input order for stable membership
        crit = assigned.get(cat, _UNASSIGNED_NAME)
        groups.setdefault(crit, []).append(cat)

    ordered_names = [c for c in declared if c in groups]
    ordered_names += [c for c in order_extra if c in groups and c not in ordered_names]
    # Any remaining criteria keys (e.g. the unassigned bucket) at the end.
    for c in groups:
        if c not in ordered_names and c != _UNASSIGNED_NAME:
            ordered_names.append(c)
    if _UNASSIGNED_NAME in groups:
        ordered_names.append(_UNASSIGNED_NAME)

    return {
        "groups": [
            {"name": name, "categories": groups[name]}
            for name in ordered_names
            if groups.get(name)
        ]
    }


def _build_criteria_prompt(
    categories: list[dict], lang: str, target_groups: int = 5
) -> tuple[str, str]:
    """Build (system, user) for the criteria-grouping call.

    categories: [{"name", "review_count", "vendor_count"}]
    """
    lang_label = {"ko": "Korean", "en": "English"}.get(lang, "Korean")
    names = [str(c.get("name") or "").strip() for c in categories if (c.get("name") or "").strip()]
    system = (
        f"You group product-feature CATEGORIES into a smaller set of\n"
        f"higher-level COMPETITIVE CRITERIA — the strategic capability\n"
        f"buckets a product manager would use to compare competitors\n"
        f"(e.g. '코칭/Action 제안 프로그램' might group '동기부여·습관 형성',\n"
        f"'개인화 코칭·인사이트', '게임화·보상', '커뮤니티·소셜').\n\n"
        f"GOAL\n"
        f"  1. Discover ~{target_groups} competitive criteria that group the\n"
        f"     {len(names)} input categories by the STRATEGIC CAPABILITY they\n"
        f"     represent. Aim for {max(3, target_groups - 2)}–"
        f"{target_groups + 1} groups.\n"
        f"  2. Assign EVERY input category to exactly ONE criterion.\n"
        f"  3. Criteria names: 4–24 character noun phrases in {lang_label},\n"
        f"     VALUE-NEUTRAL (describe the capability, not good/bad). Use\n"
        f"     '/' or '·' to join related ideas when natural.\n\n"
        f"OUTPUT — JSON only, no prose, no markdown fences:\n"
        f"{{\n"
        f'  "criteria": ["criterion A", "criterion B", ...],\n'
        f'  "assignments": [\n'
        f'    {{"category": "<exact input category name>", "criterion": "criterion A"}},\n'
        f"    ...\n"
        f"  ]\n"
        f"}}\n\n"
        f"`assignments` MUST cover all {len(names)} input categories exactly\n"
        f"once. Every `category` MUST be copied verbatim from the input list.\n"
        f"Every `criterion` MUST appear in `criteria`."
    )
    lines = []
    for c in categories:
        nm = str(c.get("name") or "").strip()
        if not nm:
            continue
        rc = int(c.get("review_count") or 0)
        vc = int(c.get("vendor_count") or 0)
        lines.append(f"- {nm} | reviews={rc} | vendors={vc}")
    user_msg = "Categories to group:\n" + "\n".join(lines)
    return system, user_msg


async def suggest_criteria_with_llm(
    categories: list[dict],
    *,
    lang: str = "ko",
    model: Optional[str] = None,
    target_groups: int = 5,
) -> dict:
    """Run a Claude call that groups the v3 AI categories into ~target_groups
    competitive criteria. Returns the canonical
    {"groups": [{"name", "categories": [...]}], "_model_used": str} shape.

    categories: [{"name", "review_count", "vendor_count"}] — extra keys
    ignored. Cached by category-name set + lang + model + prompt version.
    """
    category_names = [
        str(c.get("name") or "").strip()
        for c in (categories or [])
        if isinstance(c, dict) and (c.get("name") or "").strip()
    ]
    if not category_names:
        return {"groups": [], "_model_used": None}

    if not settings.ANTHROPIC_API_KEY:
        raise RuntimeError("ANTHROPIC_API_KEY is not set")

    chosen_model = (model or settings.ANTHROPIC_MODEL).strip()
    if chosen_model not in settings.allowed_models:
        chosen_model = settings.ANTHROPIC_MODEL

    cache_key = _criteria_cache_key(category_names, lang, chosen_model)
    cached = _criteria_cache.get(cache_key)
    if cached and time.time() - cached[0] < _CRITERIA_CACHE_TTL_SECONDS:
        out = dict(cached[1])
        out["_model_used"] = chosen_model
        return out

    system, user_msg = _build_criteria_prompt(categories, lang, target_groups)
    from anthropic import AsyncAnthropic
    client = AsyncAnthropic(api_key=settings.ANTHROPIC_API_KEY)
    resp = await client.messages.create(
        model=chosen_model,
        max_tokens=4096,
        temperature=0.0,
        system=system,
        messages=[{"role": "user", "content": user_msg}],
    )
    text = _strip_fences("".join(getattr(b, "text", "") for b in resp.content))
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError as e:
        raise RuntimeError(
            f"criteria LLM returned invalid JSON: {e}; raw={text[:200]!r}"
        )
    result = _normalize_criteria(parsed, category_names)
    _criteria_cache[cache_key] = (time.time(), result)
    out = dict(result)
    out["_model_used"] = chosen_model
    return out


# ----------------------------------------------------------------------------
# Click-a-reason → fetch real reviews
# ----------------------------------------------------------------------------


def _reason_norm(s: str) -> str:
    """Normalize a reason text for fuzzy equality. Strips whitespace +
    lowercases, since CSV reasons cell joins them as 'text(N)' while
    the saved card keeps just the text — the count suffix is stripped
    upstream but normalization is defensive against odd whitespace."""
    return re.sub(r"\s+", " ", (s or "").strip().lower())


async def reason_reviews(
    session: AsyncSession,
    *,
    vendor_key: str,
    category_name: str,
    band: str,
    reason_text: str,
) -> dict:
    """Find the VendorReasonCard matching (vendor_key, category_name,
    band) and look up the saved reason whose text matches reason_text.
    Return the hydrated review records from that reason's review_ids.

    Returns:
      {
        "vendor_key": str, "category_name": str, "band": str,
        "reason_text": str, "matched_reason": bool,
        "card_id": int|None,
        "reviews": [{id, text, rating, sentiment, posted_at, ...}, ...],
        "error": str|None,
      }
    """
    band = (band or "positive").strip().lower()
    if band not in ("positive", "negative"):
        band = "positive"
    target = _reason_norm(reason_text)
    out: dict[str, Any] = {
        "vendor_key": vendor_key,
        "category_name": category_name,
        "band": band,
        "reason_text": reason_text,
        "matched_reason": False,
        "card_id": None,
        "reviews": [],
        "error": None,
    }
    cat_lower = (category_name or "").strip().lower()

    # Find candidate saved cards. Match (vendor_key, category_name, band).
    cards = (
        await session.execute(
            select(VendorReasonCard)
            .where(VendorReasonCard.vendor_key == vendor_key)
            .where(VendorReasonCard.band == band)
            .where(VendorReasonCard.hidden.is_(False))
        )
    ).scalars().all()
    matched_card = None
    for c in cards:
        if (c.category_name or "").strip().lower() == cat_lower:
            matched_card = c
            break
    if not matched_card:
        out["error"] = "no_saved_card_for_this_strength"
        return out
    out["card_id"] = matched_card.id

    # Find the matching reason inside the card's reasons payload.
    payload = matched_card.reasons
    if isinstance(payload, dict):
        reason_list = payload.get("reasons") or []
    elif isinstance(payload, list):
        reason_list = payload
    else:
        reason_list = []
    matched_entry: Optional[dict] = None
    for r in reason_list:
        if not isinstance(r, dict):
            continue
        if _reason_norm(r.get("reason") or "") == target:
            matched_entry = r
            break
    if not matched_entry:
        out["error"] = "no_matching_reason_in_saved_card"
        return out
    out["matched_reason"] = True

    review_ids = matched_entry.get("review_ids") or []
    if not isinstance(review_ids, list):
        review_ids = []
    review_ids = [int(x) for x in review_ids if isinstance(x, (int, str)) and str(x).isdigit()]
    if not review_ids:
        out["error"] = "no_review_ids_on_saved_reason"
        return out

    # Hydrate from DB.
    if len(review_ids) > 500:
        review_ids = review_ids[:500]
    rows = (
        await session.execute(
            select(
                Review.id, Review.text, Review.rating, Review.author,
                Review.posted_at, Review.collected_at,
                Analysis.sentiment, Analysis.sentiment_score,
            )
            .join(Analysis, Analysis.review_id == Review.id, isouter=True)
            .where(Review.id.in_(review_ids))
        )
    ).all()
    by_id: dict[int, dict] = {}
    for rid, text, rating, author, posted, collected, sent, sscore in rows:
        s_key = sent.value if hasattr(sent, "value") else (str(sent) if sent else None)
        by_id[int(rid)] = {
            "id": int(rid),
            "text": text or "",
            "rating": int(rating) if rating is not None else None,
            "author": author or "",
            "posted_at": posted.isoformat() if posted else None,
            "collected_at": collected.isoformat() if collected else None,
            "sentiment": s_key,
            "sentiment_score": int(sscore) if sscore is not None else None,
        }
    out["reviews"] = [by_id[i] for i in review_ids if i in by_id]
    return out


async def criterion_reviews(
    session: AsyncSession,
    *,
    descriptors: list[dict],
    max_reviews: int = 2000,
) -> dict:
    """Aggregate reviews for a whole competitive criterion, grouped by
    the criterion's member CATEGORY and then by VENDOR.

    A criterion bundles several v3 AI categories; each category has, per
    vendor, a set of reasons. `descriptors` is the flattened list of
    those reasons (collected client-side from the per-category drill-down
    buttons), each:
      {top_category, vendor_key, vendor_display, category_name, band, reason_text}
    where `top_category` is the v3 AI category (criterion member) the
    reason was assigned to, and `category_name` is the vendor's own
    strength/weakness category used to chase the saved VendorReasonCard.

    For every distinct (vendor_key, category_name, band) we chase the
    saved VendorReasonCard (same lookup as reason_reviews), match the
    requested reason texts inside it, then bucket the review_ids by
    (top_category, vendor) — deduped within each bucket. One hydrate
    query at the end.

    Returns:
      {
        "categories": [
          {
            "name": str,                 # the AI category (criterion member)
            "review_count": int,         # displayed reviews in this category
            "vendors": [
              {key, display, review_count, reviews: [ {id, text, ...}, ... ]},
              ...                        # sorted by review_count desc
            ]
          },
          ...                            # in the order they first appear
        ],
        "total_reviews": int,            # distinct reviews hydrated
        "matched_reasons": int,
        "missing_cards": int,            # (vendor,cat,band) with no saved card
      }
    """
    from collections import defaultdict

    # Group reasons by the VendorReasonCard lookup key; remember each
    # reason's owning top_category so results can be re-grouped.
    card_groups: dict[tuple[str, str, str], list[tuple[str, str]]] = defaultdict(list)
    vendor_display: dict[str, str] = {}
    top_order: list[str] = []
    top_seen: set[str] = set()
    for d in descriptors or []:
        if not isinstance(d, dict):
            continue
        vk = (d.get("vendor_key") or "").strip()
        cat = (d.get("category_name") or "").strip()
        band = (d.get("band") or "positive").strip().lower()
        if band not in ("positive", "negative"):
            band = "positive"
        rt = (d.get("reason_text") or "").strip()
        if not vk or not cat or not rt:
            continue
        top = (d.get("top_category") or "").strip() or cat
        card_groups[(vk, cat.lower(), band)].append((rt, top))
        if vk not in vendor_display and (d.get("vendor_display") or "").strip():
            vendor_display[vk] = d["vendor_display"].strip()
        if top not in top_seen:
            top_seen.add(top)
            top_order.append(top)

    # (top_category, vendor_key) → ordered unique review ids.
    bucket_ids: dict[tuple[str, str], list[int]] = defaultdict(list)
    bucket_seen: dict[tuple[str, str], set[int]] = defaultdict(set)
    matched_reasons = 0
    missing_cards = 0

    for (vk, cat_lower, band), items in card_groups.items():
        cards = (
            await session.execute(
                select(VendorReasonCard)
                .where(VendorReasonCard.vendor_key == vk)
                .where(VendorReasonCard.band == band)
                .where(VendorReasonCard.hidden.is_(False))
            )
        ).scalars().all()
        card = None
        for c in cards:
            if (c.category_name or "").strip().lower() == cat_lower:
                card = c
                break
        if not card:
            missing_cards += 1
            continue
        payload = card.reasons
        if isinstance(payload, dict):
            reason_list = payload.get("reasons") or []
        elif isinstance(payload, list):
            reason_list = payload
        else:
            reason_list = []
        idx: dict[str, dict] = {}
        for r in reason_list:
            if isinstance(r, dict):
                idx.setdefault(_reason_norm(r.get("reason") or ""), r)

        for rt, top in items:
            entry = idx.get(_reason_norm(rt))
            if not entry:
                continue
            matched_reasons += 1
            key = (top, vk)
            seen = bucket_seen[key]
            ids = bucket_ids[key]
            for x in (entry.get("review_ids") or []):
                if isinstance(x, (int, str)) and str(x).isdigit():
                    xi = int(x)
                    if xi not in seen:
                        seen.add(xi)
                        ids.append(xi)

    # Hydrate every collected id in one query (capped).
    all_ids: list[int] = []
    global_seen: set[int] = set()
    for ids in bucket_ids.values():
        for i in ids:
            if i not in global_seen:
                global_seen.add(i)
                all_ids.append(i)
    if len(all_ids) > max_reviews:
        all_ids = all_ids[:max_reviews]
    cap_set = set(all_ids)

    by_id: dict[int, dict] = {}
    if all_ids:
        rows = (
            await session.execute(
                select(
                    Review.id, Review.text, Review.rating, Review.author,
                    Review.posted_at, Review.collected_at,
                    Analysis.sentiment, Analysis.sentiment_score,
                )
                .join(Analysis, Analysis.review_id == Review.id, isouter=True)
                .where(Review.id.in_(all_ids))
            )
        ).all()
        for rid, text, rating, author, posted, collected, sent, sscore in rows:
            s_key = sent.value if hasattr(sent, "value") else (str(sent) if sent else None)
            by_id[int(rid)] = {
                "id": int(rid),
                "text": text or "",
                "rating": int(rating) if rating is not None else None,
                "author": author or "",
                "posted_at": posted.isoformat() if posted else None,
                "collected_at": collected.isoformat() if collected else None,
                "sentiment": s_key,
                "sentiment_score": int(sscore) if sscore is not None else None,
            }

    # Assemble categories → vendors in first-seen category order.
    cats_acc: dict[str, dict[str, list[int]]] = defaultdict(dict)
    for (top, vk), ids in bucket_ids.items():
        cats_acc[top][vk] = ids
    categories_out: list[dict] = []
    for top in top_order:
        vmap = cats_acc.get(top)
        if not vmap:
            continue
        vendors: list[dict] = []
        for vk, ids in vmap.items():
            revs = [by_id[i] for i in ids if i in cap_set and i in by_id]
            if not revs:
                continue
            vendors.append({
                "key": vk,
                "display": vendor_display.get(vk, vk),
                "review_count": len(revs),
                "reviews": revs,
            })
        if not vendors:
            continue
        vendors.sort(key=lambda v: -v["review_count"])
        categories_out.append({
            "name": top,
            "review_count": sum(v["review_count"] for v in vendors),
            "vendors": vendors,
        })

    return {
        "categories": categories_out,
        "total_reviews": len(by_id),
        "matched_reasons": matched_reasons,
        "missing_cards": missing_cards,
    }


# ----------------------------------------------------------------------------
# Saved-card CRUD (mirrors the v2 pattern)
# ----------------------------------------------------------------------------


async def list_v3_cards(
    session: AsyncSession, *, include_hidden: bool = False
) -> list[dict]:
    stmt = select(CompetitiveV3Card)
    if not include_hidden:
        stmt = stmt.where(CompetitiveV3Card.hidden.is_(False))
    stmt = stmt.order_by(CompetitiveV3Card.updated_at.desc())
    rows = (await session.execute(stmt)).scalars().all()
    return [
        {
            "id": c.id,
            "label": c.label,
            "model_used": c.model_used,
            "input_filename": c.input_filename,
            "hidden": c.hidden,
            "categories_count": len((c.result_payload or {}).get("categories") or []),
            "criteria_count": len((c.criteria_mapping or {}).get("groups") or []),
            "totals": (c.result_payload or {}).get("totals") or {},
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "updated_at": c.updated_at.isoformat() if c.updated_at else None,
        }
        for c in rows
    ]


async def get_v3_card(session: AsyncSession, card_id: int) -> Optional[CompetitiveV3Card]:
    return await session.get(CompetitiveV3Card, card_id)


async def save_v3_card(
    session: AsyncSession,
    *,
    label: str,
    rows: list[dict],
    result: dict,
    model_used: Optional[str] = None,
    input_filename: Optional[str] = None,
    criteria_mapping: Optional[dict] = None,
) -> CompetitiveV3Card:
    label = (label or "").strip()
    if not label:
        raise ValueError("label cannot be empty")
    card = CompetitiveV3Card(
        label=label[:200],
        model_used=(model_used or "")[:100] or None,
        input_filename=(input_filename or "")[:255] or None,
        input_rows=list(rows or []),
        result_payload=dict(result or {}),
        criteria_mapping=dict(criteria_mapping) if criteria_mapping else None,
        hidden=False,
    )
    session.add(card)
    await session.commit()
    await session.refresh(card)
    return card


async def update_v3_card_label(
    session: AsyncSession, card_id: int, label: str
) -> Optional[CompetitiveV3Card]:
    card = await session.get(CompetitiveV3Card, card_id)
    if not card:
        return None
    label = (label or "").strip()
    if not label:
        raise ValueError("label cannot be empty")
    card.label = label[:200]
    await session.commit()
    await session.refresh(card)
    return card


async def update_v3_card_criteria(
    session: AsyncSession, card_id: int, criteria_mapping: Optional[dict]
) -> Optional[CompetitiveV3Card]:
    """Persist (or clear) the competitive-criteria grouping on a saved
    card. Pass None / empty to clear. Normalizes to the canonical
    {"groups": [{"name", "categories": [...]}]} shape, dropping junk."""
    card = await session.get(CompetitiveV3Card, card_id)
    if not card:
        return None
    if not criteria_mapping or not (criteria_mapping.get("groups")):
        card.criteria_mapping = None
    else:
        clean_groups = []
        for g in criteria_mapping.get("groups") or []:
            if not isinstance(g, dict):
                continue
            name = str(g.get("name") or "").strip()
            cats = [
                str(c).strip()
                for c in (g.get("categories") or [])
                if str(c).strip()
            ]
            if name and cats:
                clean_groups.append({"name": name[:200], "categories": cats})
        card.criteria_mapping = {"groups": clean_groups} if clean_groups else None
    await session.commit()
    await session.refresh(card)
    return card


async def toggle_v3_card_hidden(
    session: AsyncSession, card_id: int, hidden: bool
) -> Optional[CompetitiveV3Card]:
    card = await session.get(CompetitiveV3Card, card_id)
    if not card:
        return None
    card.hidden = bool(hidden)
    await session.commit()
    await session.refresh(card)
    return card


async def delete_v3_card(session: AsyncSession, card_id: int) -> bool:
    card = await session.get(CompetitiveV3Card, card_id)
    if not card:
        return False
    await session.delete(card)
    await session.commit()
    return True


# ----------------------------------------------------------------------------
# XLSX export — mirrors the user's vendor_analysis_categorized.xlsx
# ----------------------------------------------------------------------------


def export_categorized_xlsx(rows: list[dict], result: dict, *, title: Optional[str] = None) -> bytes:
    """Build a two-sheet xlsx from a categorized rows + view payload:

      Sheet "summary":
        - Title row (bold)
        - "카테고리 분포 (리뷰 수 기준)" section: category × distribution
        - "벤더 × 카테고리 (리뷰 수 합계)" section: vendor × category matrix
      Sheet "vendor_analysis":
        - Same per-reason rows the user uploaded, with the 카테고리
          column appended. Matches the format of the user's reference
          file (vendor_analysis_categorized.xlsx).

    Returns the workbook bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # Pastel palette for per-category row fills. Light enough that
    # black cell text stays readable against any of these; distinct
    # enough that scanning the sheet, the eye picks out category
    # blocks at a glance. ~15 colors so 10–12 category exports stay
    # within unique fills, with a graceful wrap on the unusual >15 case.
    _CATEGORY_PALETTE = [
        "FFE4E1",  # misty rose
        "DCEFFF",  # baby blue
        "E2F5D8",  # tea green
        "FFF1D6",  # cream / peach
        "ECDDFC",  # lavender
        "FFE0F0",  # pink lace
        "D8F5EB",  # mint
        "FFE0B5",  # apricot
        "DBE3FF",  # periwinkle
        "FFCFCF",  # blush
        "E8F8D6",  # light lime
        "E0E0FF",  # very pale indigo
        "FFFAC8",  # pale lemon
        "FCD6FF",  # pale orchid
        "D6F8FF",  # pale cyan
    ]

    def _palette_for(categories_list: list[dict]) -> dict[str, PatternFill]:
        """Map each unique category name → a PatternFill. Order is
        determined by the categories list (already sorted by review
        count desc by build_categorized_view) so the biggest category
        always gets palette[0]. Stable across runs for the same input."""
        out: dict[str, PatternFill] = {}
        idx = 0
        for c in categories_list or []:
            if not isinstance(c, dict):
                continue
            name = (c.get("name") or "").strip()
            if not name or name in out:
                continue
            color = _CATEGORY_PALETTE[idx % len(_CATEGORY_PALETTE)]
            out[name] = PatternFill("solid", fgColor=color)
            idx += 1
        return out

    wb = Workbook()

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "summary"

    totals = (result or {}).get("totals") or {}
    categories = (result or {}).get("categories") or []
    matrix = (result or {}).get("matrix") or {}

    bold = Font(bold=True)
    section_fill = PatternFill("solid", fgColor="E5E7EB")
    header_fill = PatternFill("solid", fgColor="F3F4F6")
    total_fill = PatternFill("solid", fgColor="E5E7EB")

    # Build the per-category color map once — used to tint distribution
    # rows, matrix rows, and (most importantly) the vendor_analysis
    # rows so users can scan category clusters at a glance in Excel.
    cat_fills = _palette_for(categories)

    title_text = title or (
        f"카테고리 요약 — Strength VoC "
        f"(리뷰 {totals.get('reviews', 0):,}건 / "
        f"reason {totals.get('reasons', 0)}항목 / "
        f"{totals.get('vendors', 0)}개 벤더)"
    )
    ws.cell(row=1, column=1, value=title_text).font = Font(bold=True, size=12)

    # Section 1 — category distribution
    r = 3
    ws.cell(row=r, column=1, value="카테고리 분포 (리뷰 수 기준)").font = bold
    ws.cell(row=r, column=1).fill = section_fill
    r += 1
    dist_headers = ["카테고리", "리뷰 수(합)", "비중", "reason 항목수"]
    for c_idx, h in enumerate(dist_headers, start=1):
        cell = ws.cell(row=r, column=c_idx, value=h)
        cell.font = bold
        cell.fill = header_fill
    r += 1
    for cat in categories:
        cat_name = (cat.get("name") or "").strip()
        fill = cat_fills.get(cat_name)
        ws.cell(row=r, column=1, value=cat_name)
        ws.cell(row=r, column=2, value=int(cat.get("review_count") or 0))
        ws.cell(row=r, column=3, value=(cat.get("pct") or 0) / 100.0).number_format = "0.0%"
        ws.cell(row=r, column=4, value=int(cat.get("reason_count") or 0))
        # Tint the entire distribution row so the user can pair the
        # color to vendor_analysis rows of the same category.
        if fill is not None:
            for c_idx in range(1, len(dist_headers) + 1):
                ws.cell(row=r, column=c_idx).fill = fill
        r += 1
    ws.cell(row=r, column=1, value="합계").font = bold
    ws.cell(row=r, column=2, value=int(totals.get("reviews") or 0)).font = bold
    ws.cell(row=r, column=3, value=1.0 if totals.get("reviews") else 0.0).number_format = "0.0%"
    ws.cell(row=r, column=3).font = bold
    ws.cell(row=r, column=4, value=int(totals.get("reasons") or 0)).font = bold
    for c_idx in range(1, 5):
        ws.cell(row=r, column=c_idx).fill = total_fill
    r += 2

    # Section 2 — vendor × category matrix
    ws.cell(row=r, column=1, value="벤더 × 카테고리 (리뷰 수 합계)").font = bold
    ws.cell(row=r, column=1).fill = section_fill
    r += 1
    mx_cats = matrix.get("categories") or []
    mx_vendors = matrix.get("vendors") or []
    mx_cells = matrix.get("cells") or []
    mx_row_totals = matrix.get("row_totals") or []
    mx_col_totals = matrix.get("col_totals") or []
    mx_grand_total = matrix.get("grand_total") or 0

    matrix_headers = ["카테고리 \\ 벤더"] + [v.get("display") for v in mx_vendors] + ["합계"]
    for c_idx, h in enumerate(matrix_headers, start=1):
        cell = ws.cell(row=r, column=c_idx, value=h)
        cell.font = bold
        cell.fill = header_fill
    r += 1
    for ri, cat_name in enumerate(mx_cats):
        label_cell = ws.cell(row=r, column=1, value=cat_name)
        label_cell.font = bold
        # Color only the category label cell in the matrix — the
        # numeric body cells stay uncolored so users can still read
        # them as a plain heatmap.
        fill = cat_fills.get((cat_name or "").strip())
        if fill is not None:
            label_cell.fill = fill
        for ci, n in enumerate(mx_cells[ri] if ri < len(mx_cells) else [], start=1):
            ws.cell(row=r, column=1 + ci, value=int(n or 0))
        # row total
        ws.cell(
            row=r,
            column=1 + len(mx_vendors) + 1,
            value=int(mx_row_totals[ri] or 0) if ri < len(mx_row_totals) else 0,
        ).font = bold
        r += 1
    # column totals + grand total
    ws.cell(row=r, column=1, value="합계").font = bold
    for ci, n in enumerate(mx_col_totals, start=1):
        cell = ws.cell(row=r, column=1 + ci, value=int(n or 0))
        cell.font = bold
        cell.fill = total_fill
    cell = ws.cell(row=r, column=1 + len(mx_vendors) + 1, value=int(mx_grand_total or 0))
    cell.font = bold
    cell.fill = total_fill
    ws.cell(row=r, column=1).fill = total_fill

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    for i in range(len(mx_vendors)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 18
    ws.column_dimensions[get_column_letter(2 + len(mx_vendors))].width = 12

    # ---- Vendor analysis sheet ----
    ws2 = wb.create_sheet("vendor_analysis")
    va_headers = [
        "vendor", "type", "category", "pct", "count",
        "wilson_score", "description", "small_sample",
        "reason", "review_count", "카테고리",
    ]
    for c_idx, h in enumerate(va_headers, start=1):
        cell = ws2.cell(row=1, column=c_idx, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    r = 2
    for raw in rows or []:
        if not isinstance(raw, dict):
            continue
        if not (raw.get("reason") or "").strip():
            continue
        # Coerce common columns to their natural types when possible
        try:
            pct_v = float(raw.get("pct")) if raw.get("pct") not in (None, "") else None
        except (TypeError, ValueError):
            pct_v = raw.get("pct")
        try:
            count_v = int(float(raw.get("count"))) if raw.get("count") not in (None, "") else None
        except (TypeError, ValueError):
            count_v = raw.get("count")
        try:
            review_count_v = (
                int(float(raw.get("review_count"))) if raw.get("review_count") not in (None, "") else None
            )
        except (TypeError, ValueError):
            review_count_v = raw.get("review_count")
        try:
            wilson_v = (
                float(raw.get("wilson_score")) if raw.get("wilson_score") not in (None, "") else None
            )
        except (TypeError, ValueError):
            wilson_v = raw.get("wilson_score")

        cat_name_v = (raw.get("카테고리") or "").strip()
        values = [
            raw.get("vendor") or "",
            raw.get("type") or "",
            raw.get("category") or "",
            pct_v,
            count_v,
            wilson_v,
            raw.get("description") or "",
            raw.get("small_sample") or "",
            raw.get("reason") or "",
            review_count_v,
            cat_name_v,
        ]
        row_fill = cat_fills.get(cat_name_v) if cat_name_v else None
        for c_idx, v in enumerate(values, start=1):
            cell = ws2.cell(row=r, column=c_idx, value=v)
            if row_fill is not None:
                cell.fill = row_fill
        r += 1

    # Column widths
    va_widths = {
        "vendor": 28, "type": 10, "category": 32, "pct": 8, "count": 8,
        "wilson_score": 12, "description": 40, "small_sample": 12,
        "reason": 50, "review_count": 14, "카테고리": 28,
    }
    for c_idx, h in enumerate(va_headers, start=1):
        ws2.column_dimensions[get_column_letter(c_idx)].width = va_widths.get(h, 16)
    ws2.freeze_panes = "A2"
    last_col_letter = get_column_letter(len(va_headers))
    ws2.auto_filter.ref = f"A1:{last_col_letter}{max(1, r - 1)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
